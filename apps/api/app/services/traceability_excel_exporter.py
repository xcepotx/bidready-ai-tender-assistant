from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DONE_STATUSES = {"done", "closed", "completed", "cancelled", "canceled", "resolved", "accepted", "approved"}
BAD_COMPLIANCE = {"needs_review", "needs_clarification", "partially_compliant", "non_compliant", "blocked"}

GAP_LABELS = {
    "missing_response": {"en": "Missing response", "id": "Respons kosong", "tone": "critical"},
    "missing_evidence": {"en": "Missing evidence", "id": "Evidence kosong", "tone": "warning"},
    "missing_proposal": {"en": "Missing proposal", "id": "Proposal kosong", "tone": "warning"},
    "compliance_gap": {"en": "Compliance gap", "id": "Gap compliance", "tone": "critical"},
    "open_actions": {"en": "Open actions", "id": "Action open", "tone": "warning"},
    "open_clarifications": {"en": "Open clarifications", "id": "Klarifikasi open", "tone": "warning"},
    "high_risk": {"en": "High risk", "id": "High risk", "tone": "critical"},
}


def _label(en: str, id_text: str, output_language: str) -> str:
    return id_text if str(output_language or "en").lower().startswith("id") else en


def _as_list(value: Any) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    return []


def _ids(values: Any) -> set[int]:
    result = set()
    for value in _as_list(values):
        try:
            result.add(int(value))
        except (TypeError, ValueError):
            continue
    return result


def _has_id(values: Any, item_id: int) -> bool:
    return int(item_id) in _ids(values)


def _intersects(values: Any, ids: set[int]) -> bool:
    if not ids:
        return False
    return bool(_ids(values) & ids)


def _is_done(status: Any) -> bool:
    return str(status or "").lower() in DONE_STATUSES


def _gap_label(gap: str, output_language: str) -> str:
    config = GAP_LABELS.get(gap)
    if not config:
        return gap.replace("_", " ")
    return _label(config["en"], config["id"], output_language)


def _coverage_status(gaps: list[str], output_language: str) -> str:
    if not gaps:
        return _label("Complete", "Lengkap", output_language)

    has_critical = any(GAP_LABELS.get(gap, {}).get("tone") == "critical" for gap in gaps)
    if has_critical:
        return _label("Critical gap", "Gap kritikal", output_language)

    return _label("Needs attention", "Perlu perhatian", output_language)


def _cell_value(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value) if value else "-"
    text = str(value).strip()
    return text if text else "-"


def _build_rows(
    requirements: list,
    clarifications: list,
    response_items: list,
    proposal_sections: list,
    evidence_items: list,
    compliance_items: list,
    risk_items: list,
    action_items: list,
    output_language: str,
) -> list[dict]:
    rows = []

    for req in requirements:
        req_id = int(req.id)

        responses = [item for item in response_items if int(getattr(item, "requirement_id", 0) or 0) == req_id]
        response_ids = {int(item.id) for item in responses if getattr(item, "id", None) is not None}

        evidences = [
            item for item in evidence_items
            if _has_id(getattr(item, "related_requirement_ids", None), req_id)
            or _intersects(getattr(item, "related_response_item_ids", None), response_ids)
        ]
        evidence_ids = {int(item.id) for item in evidences if getattr(item, "id", None) is not None}

        proposals = [
            item for item in proposal_sections
            if _intersects(getattr(item, "source_response_item_ids", None), response_ids)
            or _intersects(getattr(item, "related_evidence_item_ids", None), evidence_ids)
        ]
        proposal_ids = {int(item.id) for item in proposals if getattr(item, "id", None) is not None}

        req_clarifications = [
            item for item in clarifications
            if int(getattr(item, "requirement_id", 0) or 0) == req_id
        ]

        compliance = [
            item for item in compliance_items
            if int(getattr(item, "requirement_id", 0) or 0) == req_id
            or int(getattr(item, "response_item_id", 0) or 0) in response_ids
        ]

        risks = [
            item for item in risk_items
            if _has_id(getattr(item, "related_requirement_ids", None), req_id)
            or _intersects(getattr(item, "related_response_item_ids", None), response_ids)
            or _intersects(getattr(item, "related_evidence_item_ids", None), evidence_ids)
            or (getattr(item, "source_type", None) == "requirement" and int(getattr(item, "source_id", 0) or 0) == req_id)
        ]

        actions = [
            item for item in action_items
            if _has_id(getattr(item, "related_requirement_ids", None), req_id)
            or _intersects(getattr(item, "related_response_item_ids", None), response_ids)
            or _intersects(getattr(item, "related_evidence_item_ids", None), evidence_ids)
            or _intersects(getattr(item, "related_proposal_section_ids", None), proposal_ids)
            or (getattr(item, "source_type", None) == "requirement" and int(getattr(item, "source_id", 0) or 0) == req_id)
        ]

        open_actions = [item for item in actions if not _is_done(getattr(item, "status", None))]
        open_clarifications = [item for item in req_clarifications if not _is_done(getattr(item, "status", None))]
        high_risks = [
            item for item in risks
            if str(getattr(item, "severity", None) or getattr(item, "risk_level", "")).lower() in {"high", "critical"}
        ]

        compliance_gap = any(
            str(getattr(item, "compliance_status", "") or "").lower() in BAD_COMPLIANCE
            for item in compliance
        )

        gaps = []
        if not responses:
            gaps.append("missing_response")
        if not evidences:
            gaps.append("missing_evidence")
        if not proposals:
            gaps.append("missing_proposal")
        if compliance_gap:
            gaps.append("compliance_gap")
        if open_actions:
            gaps.append("open_actions")
        if open_clarifications:
            gaps.append("open_clarifications")
        if high_risks:
            gaps.append("high_risk")

        rows.append({
            "requirement_id": req_id,
            "category": _cell_value(getattr(req, "category", None)),
            "requirement_text": _cell_value(getattr(req, "requirement_text", None)),
            "owner": _cell_value(getattr(req, "suggested_owner", None)),
            "requirement_status": _cell_value(getattr(req, "status", None)),
            "coverage_status": _coverage_status(gaps, output_language),
            "response_count": len(responses),
            "evidence_count": len(evidences),
            "proposal_count": len(proposals),
            "compliance_status": _cell_value(getattr(compliance[0], "compliance_status", None) if compliance else None),
            "risk_count": len(risks),
            "high_risk_count": len(high_risks),
            "action_count": len(actions),
            "open_action_count": len(open_actions),
            "open_clarification_count": len(open_clarifications),
            "gaps": ", ".join(_gap_label(gap, output_language) for gap in gaps) if gaps else _label("Complete", "Lengkap", output_language),
            "gap_count": len(gaps),
        })

    status_rank = {
        _label("Critical gap", "Gap kritikal", output_language): 0,
        _label("Needs attention", "Perlu perhatian", output_language): 1,
        _label("Complete", "Lengkap", output_language): 2,
    }

    return sorted(rows, key=lambda row: (status_rank.get(row["coverage_status"], 9), -row["gap_count"], row["requirement_id"]))


def _style_sheet(ws, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="0B1D3A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value or "")
            max_length = max(max_length, min(len(value), 80))

        ws.column_dimensions[column_letter].width = max(12, min(max_length + 3, 55))


def export_traceability_matrix_xlsx(
    *,
    project,
    requirements: list,
    clarifications: list,
    response_items: list,
    proposal_sections: list,
    evidence_items: list,
    compliance_items: list,
    risk_items: list,
    action_items: list,
    output_path: str,
    output_language: str = "en",
) -> str:
    rows = _build_rows(
        requirements=requirements,
        clarifications=clarifications,
        response_items=response_items,
        proposal_sections=proposal_sections,
        evidence_items=evidence_items,
        compliance_items=compliance_items,
        risk_items=risk_items,
        action_items=action_items,
        output_language=output_language,
    )

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    labels = {
        "project": _label("Project", "Project", output_language),
        "issuer": _label("Issuer", "Issuer", output_language),
        "generated_at": _label("Generated At", "Dibuat Pada", output_language),
        "output_language": _label("Output Language", "Bahasa Output", output_language),
        "requirements": _label("Requirements", "Requirement", output_language),
        "complete": _label("Complete", "Lengkap", output_language),
        "critical": _label("Critical Gaps", "Gap Kritikal", output_language),
        "attention": _label("Needs Attention", "Perlu Perhatian", output_language),
    }

    complete_label = _label("Complete", "Lengkap", output_language)
    critical_label = _label("Critical gap", "Gap kritikal", output_language)
    attention_label = _label("Needs attention", "Perlu perhatian", output_language)

    summary_data = [
        [labels["project"], getattr(project, "title", "-")],
        [labels["issuer"], getattr(project, "issuer", "-")],
        [labels["generated_at"], datetime.now(timezone.utc).isoformat()],
        [labels["output_language"], output_language],
        [labels["requirements"], len(rows)],
        [labels["complete"], sum(1 for row in rows if row["coverage_status"] == complete_label)],
        [labels["critical"], sum(1 for row in rows if row["coverage_status"] == critical_label)],
        [labels["attention"], sum(1 for row in rows if row["coverage_status"] == attention_label)],
    ]

    for item in summary_data:
        summary_ws.append(item)

    matrix_ws = wb.create_sheet("Traceability Matrix")

    headers = [
        _label("Requirement ID", "ID Requirement", output_language),
        _label("Category", "Kategori", output_language),
        _label("Requirement Text", "Teks Requirement", output_language),
        _label("Owner", "PIC", output_language),
        _label("Requirement Status", "Status Requirement", output_language),
        _label("Coverage Status", "Status Coverage", output_language),
        _label("Response Count", "Jumlah Respons", output_language),
        _label("Evidence Count", "Jumlah Evidence", output_language),
        _label("Proposal Count", "Jumlah Proposal", output_language),
        _label("Compliance Status", "Status Compliance", output_language),
        _label("Risk Count", "Jumlah Risk", output_language),
        _label("High Risk Count", "Jumlah High Risk", output_language),
        _label("Action Count", "Jumlah Action", output_language),
        _label("Open Action Count", "Jumlah Action Open", output_language),
        _label("Open Clarification Count", "Jumlah Klarifikasi Open", output_language),
        _label("Gaps", "Gap", output_language),
    ]

    matrix_ws.append(headers)

    for row in rows:
        matrix_ws.append([
            row["requirement_id"],
            row["category"],
            row["requirement_text"],
            row["owner"],
            row["requirement_status"],
            row["coverage_status"],
            row["response_count"],
            row["evidence_count"],
            row["proposal_count"],
            row["compliance_status"],
            row["risk_count"],
            row["high_risk_count"],
            row["action_count"],
            row["open_action_count"],
            row["open_clarification_count"],
            row["gaps"],
        ])

    _style_sheet(summary_ws, freeze="A2")
    _style_sheet(matrix_ws, freeze="A2")

    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 56
    matrix_ws.column_dimensions["C"].width = 70
    matrix_ws.column_dimensions["P"].width = 45

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return str(output)
